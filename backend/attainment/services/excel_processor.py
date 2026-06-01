from openpyxl import load_workbook


def _clean(value):
    if value in (None, "#REF!"):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _number(value, default=0):
    value = _clean(value)
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return default


def _normalize_co(value):
    """Normalize CO string: fix C0x (zero) -> COx, uppercase, strip spaces."""
    import re
    # Replace C followed by zero(s) then digit(s) -> CO + digits, e.g. C02 -> CO2
    return re.sub(r'\bC0+(\d+)\b', r'CO\1', str(value).strip().upper())


def _is_co(value):
    """Accept single CO (CO1) or multi-CO (CO1,CO2 / CO1/CO2). Also handles C02 typo."""
    if not isinstance(value, str):
        return False
    normalized = _normalize_co(value)
    parts = [p.strip() for p in normalized.replace("/", ",").split(",") if p.strip()]
    return bool(parts) and all(p.startswith("CO") and p[2:].isdigit() for p in parts)


def _parse_cos(value):
    """Return list of normalized CO ids from a cell value like 'CO1', 'CO1,CO2', 'C02/CO1'."""
    normalized = _normalize_co(str(value))
    parts = [p.strip() for p in normalized.replace("/", ",").split(",") if p.strip()]
    return [p for p in parts if p.startswith("CO") and p[2:].isdigit()]


def _assignment_weight(question_id):
    return 0.75 if str(question_id).strip().upper().startswith("A") else 1


def _text(value, default=""):
    value = _clean(value)
    return str(value) if value is not None else default


def _sheet(wb, name):
    return wb[name] if name in wb.sheetnames else None


def _profile(front_page):
    fields = {
        "courseName": "D2",
        "courseCode": "D3",
        "academicYear": "D4",
        "semester": "D5",
        "programme": "D6",
        "specialization": "D7",
        "courseYear": "D8",
        "courseSemester": "D9",
        "credits": "D10",
        "faculty": "D11",
    }
    profile = {}
    if not front_page:
        return profile
    for key, cell in fields.items():
        profile[key] = _text(front_page[cell].value)
    return profile


def _evaluation_policy(front_page):
    if not front_page:
        return {}
    return {
        "interimTest": _number(front_page["D21"].value),
        "endExam": _number(front_page["D22"].value),
        "continuousEvaluation": _number(front_page["D23"].value),
        "other": _number(front_page["D24"].value),
        "total": _number(front_page["D25"].value),
    }


def _grading_policy(front_page):
    if not front_page:
        return []
    policy = []
    for row in range(29, 36):
        grade = _text(front_page[f"C{row}"].value)
        if grade:
            policy.append(
                {
                    "grade": grade,
                    "lower": _number(front_page[f"D{row}"].value),
                    "upper": _number(front_page[f"E{row}"].value),
                }
            )
    return policy


def _mapping(front_page, fallback_sheet=None):
    source = front_page or fallback_sheet
    if not source:
        return {}, []

    header_row = 13 if source.title == "Front page" else 19
    co_start = 14 if source.title == "Front page" else 20
    co_column = 7 if source.title == "Front page" else 2
    po_start = 8 if source.title == "Front page" else 3

    pos = []
    for column in range(po_start, po_start + 16):
        po = _text(source.cell(header_row, column).value)
        if po and (po.startswith("PO") or po.startswith("PSO")):
            pos.append(po)

    mapping = {}
    for row in range(co_start, co_start + 8):
        co = _text(source.cell(row, co_column).value).replace(":", "")
        if not co.startswith("CO"):
            continue
        mapping[co] = {}
        for index, po in enumerate(pos):
            mapping[co][po] = _number(source.cell(row, po_start + index).value)
    return mapping, pos


def _course_outcomes(front_page, mapping):
    cos = []
    if front_page:
        for row in range(14, 22):
            co = _text(front_page[f"C{row}"].value).replace(":", "")
            if co.startswith("CO"):
                cos.append({"id": co, "description": "", "target": _number(front_page["D32"].value, 50)})
    if not cos:
        cos = [{"id": co, "description": "", "target": 50} for co in mapping]
    return cos


def _direct_results(sheet):
    results = []
    po_scores = {}
    if not sheet:
        return results, po_scores

    for row in range(12, 18):
        co = _text(sheet[f"B{row}"].value)
        if co.startswith("CO"):
            percentage = _number(sheet[f"D{row}"].value)
            score = _number(sheet[f"E{row}"].value)
            results.append(
                {
                    "co": co,
                    "attainmentPercentage": percentage,
                    "score": score,
                    "studentsAttained": None,
                    "totalStudents": None,
                    "target": 50,
                }
            )

    for column in range(3, 19):
        po = _text(sheet.cell(27, column).value)
        score = _clean(sheet.cell(33, column).value)
        if po and score is not None:
            po_scores[po] = _number(score)

    return results, po_scores


def _indirect_results(sheet):
    co_results = []
    po_scores = {}
    scale = {"VH": 5, "H": 4, "M": 3, "L": 2, "VL": 1}
    if not sheet:
        return {"scale": scale, "coResults": co_results, "poScores": po_scores}

    labels = ["VH", "H", "M", "L", "VL"]
    for row in range(14, 19):
        co = _text(sheet[f"B{row}"].value)
        if co.startswith("CO"):
            counts = {label: _number(sheet.cell(row, 4 + index).value) for index, label in enumerate(labels)}
            co_results.append(
                {
                    "co": co,
                    "counts": counts,
                    "total": _number(sheet[f"I{row}"].value),
                    "gradingIndex": _number(sheet[f"K{row}"].value),
                    "score": _number(sheet[f"L{row}"].value),
                }
            )

    for column in range(4, 20):
        po = _text(sheet.cell(32, column).value)
        score = _clean(sheet.cell(33, column).value)
        if po and score is not None:
            po_scores[po] = _number(score)

    return {"scale": scale, "coResults": co_results, "poScores": po_scores}


def _assessment_summary(sheet):
    if not sheet:
        return {"studentCount": 0, "questions": []}

    questions = []
    for column in range(4, 49):
        co = _text(sheet.cell(4, column).value)
        question = _text(sheet.cell(5, column).value)
        max_marks = _number(sheet.cell(6, column).value)
        if co.startswith("CO") and question and max_marks:
            questions.append({"id": question, "co": co, "maxMarks": max_marks})

    student_count = 0
    for row in range(8, sheet.max_row + 1):
        serial = _clean(sheet.cell(row, 1).value)
        if isinstance(serial, (int, float)) and serial > 0:
            student_count += 1

    return {"studentCount": student_count, "questions": questions}


def parse_accreditation_workbook(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    front_page = _sheet(workbook, "Front page")
    direct_sheet = _sheet(workbook, "CO-PO Direct Attainment")
    indirect_sheet = _sheet(workbook, "CO-PO Indirect Attainment")
    assessment_sheet = _sheet(workbook, "CO Assessment")

    mapping, pos = _mapping(front_page, direct_sheet)
    direct_results, direct_po_scores = _direct_results(direct_sheet)
    indirect = _indirect_results(indirect_sheet)
    assessment_summary = _assessment_summary(assessment_sheet)

    all_po_scores = {}
    for po in pos:
        direct = direct_po_scores.get(po, 0)
        indirect_score = indirect["poScores"].get(po, 0)
        all_po_scores[po] = round((direct * 0.8) + (indirect_score * 0.2), 2) if direct or indirect_score else 0

    direct_scores = [item["score"] for item in direct_results if item["score"]]
    po_values = [value for value in all_po_scores.values() if value]

    return {
        "course": _profile(front_page),
        "cos": _course_outcomes(front_page, mapping),
        "pos": pos,
        "mapping": mapping,
        "evaluationPolicy": _evaluation_policy(front_page),
        "gradingPolicy": _grading_policy(front_page),
        "summary": {
            "totalStudents": assessment_summary["studentCount"],
            "totalCOs": len(mapping),
            "totalPOs": len(pos),
            "averageCOScore": round(sum(direct_scores) / len(direct_scores), 2) if direct_scores else 0,
            "averagePOScore": round(sum(po_values) / len(po_values), 2) if po_values else 0,
        },
        "coResults": direct_results,
        "poScores": all_po_scores or direct_po_scores,
        "directPoScores": direct_po_scores,
        "indirect": indirect,
        "assessments": [{"id": "workbook", "name": "Workbook Assessment", "questions": assessment_summary["questions"]}],
        "sourceSheets": workbook.sheetnames,
    }


def parse_marks_workbook(uploaded_file):
    """
    Parses IA / ESE / CA marks workbooks.

    Multi-CO handling:
      A column whose CO cell contains 'CO1, CO2, CO3, CO4' (or slash-separated)
      is treated as a single question whose marks are split equally across all
      listed COs.  The storage key for each split is '<question_id>__<CO_ID>'
      so duplicate question IDs across different columns never collide.

    CA format support:
      Accepts sheets where a single question column is tagged to multiple COs
      (e.g. Quiz1 -> CO1,CO2,CO3,CO4) with one mark per student.
    """
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = None
    co_row = None
    best_score = 0

    for candidate in workbook.worksheets:
        for row in range(1, min(candidate.max_row, 20) + 1):
            # A cell counts as a CO cell if it contains at least one valid CO id
            co_count = sum(
                1 for col in range(1, candidate.max_column + 1)
                if _is_co(_text(candidate.cell(row, col).value))
            )
            max_count = sum(
                1 for col in range(1, candidate.max_column + 1)
                if _number(candidate.cell(row + 2, col).value) > 0
            )
            data_count = sum(
                1 for data_row in range(row + 3, min(candidate.max_row, row + 40) + 1)
                if isinstance(_clean(candidate.cell(data_row, 1).value), (int, float))
            )
            score = co_count + max_count + data_count
            # Relax threshold: allow co_count >= 1 to support CA sheets with one multi-CO column
            if co_count >= 1 and max_count >= 1 and data_count >= 2 and score > best_score:
                sheet = candidate
                co_row = row
                best_score = score

    if not co_row:
        raise ValueError("Could not find the CO header row in the marks sheet")

    question_row = co_row + 1
    max_row = co_row + 2

    # ------------------------------------------------------------------ #
    # Build question list.                                                 #
    # Each physical column may map to 1..N COs.  We create one entry per  #
    # (column, co) pair.  The storage key is '<question_id>__<CO_ID>'     #
    # which guarantees uniqueness even when the same label appears in      #
    # multiple columns (e.g. T1Q1A and T1Q1B both named "T1Q1").          #
    # ------------------------------------------------------------------ #
    questions = []
    seen_col_ids = {}   # question_id -> occurrence count (for same-label dedup)

    for column in range(1, sheet.max_column + 1):
        co_raw = _text(sheet.cell(co_row, column).value)
        question_id = _text(sheet.cell(question_row, column).value)
        max_marks = _number(sheet.cell(max_row, column).value)

        if not _is_co(co_raw) or not question_id or question_id.lower() == "total" or max_marks <= 0:
            continue

        co_ids_for_q = _parse_cos(co_raw)   # e.g. ['CO1','CO2','CO3','CO4']
        split = len(co_ids_for_q) or 1
        split_max = round(max_marks / split, 4)
        weight = _assignment_weight(question_id)

        # Deduplicate same question_id across different columns
        seen_col_ids[question_id] = seen_col_ids.get(question_id, 0) + 1
        occurrence = seen_col_ids[question_id]
        col_base = question_id if occurrence == 1 else f"{question_id}_x{occurrence}"

        for co in co_ids_for_q:
            # Unique storage key: always includes CO suffix when multi-CO
            uid = f"{col_base}__{co}" if split > 1 else col_base
            questions.append({
                "id": uid,           # unique key used in marks dict
                "label": question_id,  # display label (original)
                "co": co,
                "maxMarks": split_max,
                "rawMaxMarks": max_marks,
                "splitCount": split,
                "weight": weight,
                "column": column,    # physical column to read raw value from
                "colBase": col_base, # base key for raw mark lookup
            })

    if not questions:
        raise ValueError("No question columns were found below the CO header row")

    # Find first data row
    data_start = max_row + 1
    for row in range(max_row + 1, min(sheet.max_row, max_row + 8) + 1):
        if isinstance(sheet.cell(row, 1).value, (int, float)) and sheet.cell(row, 1).value > 0:
            data_start = row
            break

    # ------------------------------------------------------------------ #
    # Read student rows.                                                   #
    # raw_marks stores the split value under the uid key so the           #
    # calculator can look it up directly.                                  #
    # ------------------------------------------------------------------ #
    students = []
    unique_cols = {}   # column -> colBase (to avoid reading same column twice)
    for q in questions:
        unique_cols[q["column"]] = q["colBase"]

    for row in range(data_start, sheet.max_row + 1):
        serial = _clean(sheet.cell(row, 1).value)
        if not isinstance(serial, (int, float)):
            continue
        # Skip rows where all question columns are empty
        if all(_clean(sheet.cell(row, col).value) is None for col in unique_cols):
            continue

        # Read raw value per column once
        col_raw = {col: _number(sheet.cell(row, col).value) for col in unique_cols}

        raw_marks = {}
        for q in questions:
            raw_val = col_raw[q["column"]]
            # Split equally across COs
            raw_marks[q["id"]] = round(raw_val / q["splitCount"], 4)

        students.append({
            "registerNumber": _text(sheet.cell(row, 2).value, f"REG{int(serial):03d}"),
            "name": _text(sheet.cell(row, 3).value, f"Student {int(serial)}"),
            "section": "",
            "marks": dict(raw_marks),
            "rawMarks": raw_marks,
        })

    # ------------------------------------------------------------------ #
    # Build CO summary (raw marks, before weightage scaling)              #
    # ------------------------------------------------------------------ #
    co_ids_sorted = sorted(
        {q["co"] for q in questions},
        key=lambda v: int(v[2:]) if v[2:].isdigit() else v
    )
    co_summary = []
    for co_id in co_ids_sorted:
        co_qs = [q for q in questions if q["co"] == co_id]
        total_marks = round(sum(q["maxMarks"] for q in co_qs), 2)
        rows = []
        for student in students:
            attained = round(sum(student["rawMarks"].get(q["id"], 0) for q in co_qs), 2)
            rows.append({
                "registerNumber": student["registerNumber"],
                "name": student["name"],
                "totalMarks": total_marks,
                "marksAttained": attained,
            })
        co_summary.append({"co": co_id, "totalMarks": total_marks, "rows": rows})

    # Strip internal-only keys before returning
    export_questions = [
        {k: v for k, v in q.items() if k not in ("column", "colBase")}
        for q in questions
    ]

    return {
        "questions": export_questions,
        "students": students,
        "coSummary": co_summary,
        "message": f"{len(students)} students and {len(questions)} question columns imported",
    }
